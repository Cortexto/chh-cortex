#!/usr/bin/env python3
"""Fill LSRI HIIA budget template at $350,000 (2026-07-14)."""

from pathlib import Path
import shutil

import openpyxl

SRC = Path("/Users/berktugkubuk/Downloads/LSRI High-Impact Individual Awards Budget.xlsx")
OUT = Path(__file__).resolve().parent / "LSRI-HIIA-Budget-2026-07-14.xlsx"


def fill_budget() -> None:
    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    ws = wb["LSRI HIIA Budget"]

    # Applicant / primary contact (merged cells C9:C10 and G9:I10)
    ws["C9"] = "Shatha Elnakib"
    ws["G9"] = (
        "Shatha Elnakib, PhD, MPH | Assistant Research Professor, "
        "Center for Humanitarian Health, Johns Hopkins Bloomberg School of Public Health | "
        "selnaki1@jhu.edu"
    )

    # Key Personnel row 16 — Shatha Elnakib, PI
    ws["B16"] = "Shatha Elnakib"
    ws["C16"] = (
        "Lead PI: study design, ethics, validation framework, Türkiye field "
        "supervision, publication lead"
    )
    ws["D16"] = 150000  # verify with Shatha/JHURA
    ws["F16"] = 0.20
    ws["G16"] = 0.20

    # Postdoc Stipend row 37 — Aral Surmeli, DrPH trainee
    ws["B37"] = "Aral Surmeli"
    ws["C37"] = (
        "DrPH trainee: data-provenance mapping, audit documentation, "
        "dissertation fieldwork (trainee beneficiary, not co-PI)"
    )
    ws["D37"] = 32000
    ws["F37"] = 1.0
    ws["G37"] = 1.0

    # Postdoc Salary row 32 — CHH research analyst
    ws["B32"] = "CHH Research Analyst"
    ws["C32"] = (
        "NLP, statistics, annotation oversight, data governance under PI"
    )
    ws["D32"] = 85000
    ws["F32"] = 0.25
    ws["G32"] = 0.25

    supplies = [
        (
            "JHU secure compute / cloud GPU credits",
            5500,
            1,
            1,
        ),
        (
            "Annotation platform and NLP software licenses",
            2750,
            1,
            1,
        ),
        (
            "App/server staging, API monitoring, encrypted backup",
            3500,
            1,
            1,
        ),
        (
            "Participant incentives (cognitive, FGD, IDI)",
            2000,
            2,
            1,
        ),
        (
            "Recording, transcription, Arabic/Turkish translation",
            2500,
            1,
            1,
        ),
        (
            "Publication and open-science fees",
            1353,
            1,
            1,
        ),
    ]
    for i, (desc, cost, q1, q2) in enumerate(supplies, start=47):
        ws[f"B{i}"] = desc
        ws[f"E{i}"] = cost
        ws[f"F{i}"] = q1
        ws[f"G{i}"] = q2

    travel = [
        (
            12000,
            "Baltimore-Istanbul field validation (2 travelers: PI + trainee)",
            3,
            2,
        ),
        (
            3000,
            "Domestic CHH/JHU coordination and IRB meetings",
            1,
            1,
        ),
        (
            6000,
            "Conference dissemination (international, 2 travelers)",
            0,
            1,
        ),
    ]
    for i, (cost, desc, q1, q2) in enumerate(travel, start=61):
        ws[f"B{i}"] = cost  # Cost Per Trip (template formula = B*F, B*G)
        ws[f"C{i}"] = desc  # Description of Trip
        ws[f"F{i}"] = q1
        ws[f"G{i}"] = q2

    # HERA subcontract row 73
    ws["A73"] = "HERA Digital Health"
    ws["B73"] = (
        "De-identified data export, platform pseudonymization support, "
        "bounded API/integration for research layer (≤10% cap; no field ops)"
    )
    ws["H73"] = 17500
    ws["I73"] = 17500

    wb.save(OUT)

    wb2 = openpyxl.load_workbook(OUT, data_only=True)
    ws2 = wb2["LSRI HIIA Budget"]
    y1 = ws2["H79"].value
    y2 = ws2["I79"].value
    total = (y1 or 0) + (y2 or 0)
    print(f"Saved: {OUT}")
    print(f"Year 1 total: ${y1:,.2f}" if y1 else "Year 1: n/a")
    print(f"Year 2 total: ${y2:,.2f}" if y2 else "Year 2: n/a")
    print(f"Grand total:  ${total:,.2f}")
    if abs(total - 350000) > 1:
        print(f"WARNING: total differs from $350,000 by ${total - 350000:,.2f}")


if __name__ == "__main__":
    fill_budget()
